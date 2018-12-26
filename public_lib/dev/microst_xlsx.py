#!/usr/bin/env python
#encoding=utf-8
#Created by gaolin on 2017-10-11 10:17
import public_lib
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font, Color,PatternFill,Border,Side,Alignment

def T(*args, **kwargs):
    _str = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    menu_str = []
    for i in _str:
        menu_str.append(i)
    for i in _str:
        for l in _str:
            menu_str.append(i + l)
    return menu_str


menu_str = T()



class ExportHander: #导出数据到exlcel
    def __init__(self,**kwargs):
        self.data = kwargs['data']  #mysql查询的返回数据，格式list
        self.menu = kwargs['menu']  #excel的菜单栏[fidld,菜单名称,列宽]
        self.title = kwargs['title']
        self.merge = kwargs.get('merge',[])
        self.hieght = 28            #行高
        self.wb = Workbook()
        border_color = Side(border_style="thin", color="000000")
        self.border = Border(top=border_color, left=border_color, right=border_color, bottom=border_color)
        self.style_center = Alignment(horizontal='center')
        self.style_left = Alignment(horizontal='left')
        self.vertical_center = Alignment(vertical="center",horizontal='center')
        self.vertical_left = Alignment(vertical="center",horizontal='left')

    def _temp_data(self):
        temp_data = []
        for nu in range(10):
            p = {}
            for i in self.menu:
                p[i[0]] = ''
            temp_data.append(p)
        return temp_data

    def _merge_convert(self):
        ge = []
        if self.merge:
            for m in self.merge:
                r1 = "%s:%s" % (menu_str[m[0]]+str(m[1]+1),menu_str[m[2]]+str(m[3]+1))
                ge.append(r1)
        return ge

    def _create(self):
        sheet = self.wb.active
        sheet.title = unicode(self.title, "utf-8")
        sheet.row_dimensions[2].height = self.hieght+1
        for menu_number in range(len(self.menu)):
            menu_y = menu_str[menu_number+1]
            menu_x_y = '%s2' % menu_y
            sheet[menu_x_y] = self.menu[menu_number][1]
            sheet[menu_x_y].alignment = self.style_center
            sheet[menu_x_y].font = Font(color='FF000000',bold=True)
            sheet[menu_x_y].border = self.border
            sheet.column_dimensions[menu_y].width = self.menu[menu_number][2].get('w',16)
            # print self.menu[menu_number][1],menu_str[menu_number+1]

        for column_number in range(len(self.data)):
            info = self.data[column_number]
            for value_number in range(len(self.menu)):
                field_name = self.menu[value_number][0]
                value_data = info.get(field_name,{})
                value_x = column_number+3
                value_x_y = "%s%s" % (menu_str[value_number+1],value_x)
                try:
                    sheet[value_x_y] = value_data.get("value","")
                except  AttributeError,e:
                    sheet[value_x_y] = ""
                line_style = self.menu[value_number][2].get('style','center')
                if line_style == 'center':
                    sheet[value_x_y].alignment = self.style_center
                elif line_style == 'left':
                    sheet[value_x_y].alignment = self.style_left
                sheet[value_x_y].border = self.border
                try:
                    background_color = value_data.get("background")
                except AttributeError,e:
                    background_color = False
                if background_color:
                    sheet[value_x_y].fill = PatternFill(fgColor=background_color,fill_type = "solid")
                sheet.row_dimensions[value_x].height = self.hieght

        merge_list = self._merge_convert()
        if merge_list:
            for merge_info in merge_list:
                sheet.merge_cells(merge_info)
                merge_value = merge_info.split(':')[0]
                sheet[merge_value].alignment = self.vertical_center
                merge_style = self.menu[menu_str.index(merge_info[0])-1][2].get('style','center')
                if merge_style == 'center':
                    sheet[merge_value].alignment = self.vertical_center
                elif merge_style == 'left':
                    sheet[merge_value].alignment = self.vertical_left
    #创建第二张表
    def _create_temp(self,temp_data):
        # {"data": [['哈哈', '你好'],['哈哈', '你好'],],"title": "你好","width": [{"w":12},{"w":40,"style":"left"},{"w":12}]}
        data = temp_data['data']
        table_width = temp_data['width']
        ws1 = self.wb.create_sheet(unicode(temp_data.get('title','模板'), "utf-8"))
        for column_width in range(len(table_width)):
            ws1.column_dimensions[menu_str[column_width+1]].width = table_width[column_width]['w']

        for column_number in range(len(data)):
            info = data[column_number]
            for value_number in range(len(info)):
                value_x = column_number+2
                value_x_y = "%s%s" % (menu_str[value_number + 1], value_x)
                ws1[value_x_y]= info[value_number]
                ws1.row_dimensions[value_x].height = self.hieght
                line_style = table_width[value_number].get('style', 'center')
                ws1[value_x_y].border = self.border
                if line_style == 'center':
                    ws1[value_x_y].alignment = self.style_center
                elif line_style == 'left':
                    ws1[value_x_y].alignment = self.style_left

    def _save(self,obj,template={}):
        if template:
            self.data=self._temp_data()
            self._create_temp(template)
        self._create()
        self.wb.save(obj)

    def save(self,file_path,template={}):
        self._save(file_path,template=template)
        return file_path

    def response(self,template={}):
        import StringIO
        Buffer = StringIO.StringIO()
        self._save(Buffer,template=template)
        return Buffer

class ImportHander: #从excel读取数据
    def __init__(self,**kwargs):
        """
        read excel data
        :param kwargs:
            :primary_key:clumn data primary
            :menu: menu name,format is [[field name,menu name],[field name,menu name]]
        """
        self.data_excel = []
        self.primary_key = kwargs.get('primary_key',[])     #判断字段是否唯一,某个菜单中的数据不能重复
        self.menu = { i[1]:i[0] for i in kwargs['menu']}    #excel的菜单栏[[fidld,菜单名称],[fidld,菜单名称]]
        self.msg = {"status": False, "msg": "", "data": []}
        self.x = 2      #设置从第几列开始读取数据
        self.y = 2      #设置从第几行开始读取数据
        self.table_number = kwargs.get('table_number',0)  # 默认读第一张表
        self.menu_lenght = len(self.menu)
        self.debug = kwargs.get("debug",False)

    def _relaod(self):
        # ws1 = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])
        ws1 = self.wb[self.wb.sheetnames[self.table_number]]
        for row in range(self.y, ws1.max_row + 1):
            tmplist = []
            for column in range(self.x, ws1.max_column + 1):
                _values = ws1.cell(row=row, column=column).value
                if _values == None:
                    _values = ""
                tmplist.append(_values)
            if set(tmplist) == set([""]):
                continue
            self.data_excel.append([ str(i).encode('utf-8') for i in tmplist][:self.menu_lenght])

    def _list_data(self):
        if len(self.data_excel) < 2:
            self.msg['msg'] = "读取空数据"
        # elif max([ len(i) for i in self.data_excel])>len(self.menu):
        #     self.msg['msg'] = "第%s列,数据区域超出" % max([ len(i) for i in self.data_excel])
        if self.debug is True:
            print public_lib.json_data({"菜单名称":self.menu})
        elif len(set(self.data_excel[0]))!=len(self.menu):
            self.msg['msg'] = "菜单栏重复字段:%s" % public_lib.set_list(self.data_excel[0])
        else:
            data = []
            for number in range(len(self.data_excel[1:])):
                x = self.data_excel[1:][number]
                f = {}
                for row_number in range(len(x)):
                    excel_name = self.data_excel[0][row_number].strip().replace('\n','')
                    menu = self.menu.get(excel_name,False)
                    if menu is False:
                        self.msg['msg'] = "非法菜单名称:%s" % excel_name
                        return
                    f[menu]=x[row_number]
                data.append(f)
            if self.primary_key:
                res_menu = { m_y:m_x for m_x,m_y in self.menu.items() }
                for key_name in self.primary_key:
                    row_data = [i[key_name] for i in data]
                    if 'None' in row_data:
                        self.msg['msg'] = "%s,不得为空" % (res_menu[key_name])
                        return
                    else:
                        repeat_index = public_lib.set_list(row_data)
                        if repeat_index:
                            self.msg['msg'] = "%s,重复内容:%s" % (res_menu[key_name],repeat_index[0])
                            return
            self.msg['status'] = True
            self.msg['data'] = data

    def file(self,file_path):
        try:
            self.wb = load_workbook(filename=file_path)
        except IOError,e:
            self.msg['msg'] = '%s,文件不存在' % file_path
            return self.msg
        except Exception,e:
            self.msg['msg'] = '上传的文件必须是一个xlsx文件'
            return self.msg
        self._relaod()
        self._list_data()
        return self.msg

    def response(self,conntext=False):
        import tempfile,os
        file_path = tempfile.NamedTemporaryFile(dir=os.getcwd())
        file_path.name = file_path.name+'_mywork_tmp.xlsx'
        with open(file_path.name,r'wb+') as F:
            F.write(conntext)
        file_path.flush()
        data = self.file(file_path=file_path.name)
        file_path.close()
        return data


def excel_data_trace(data):
    _data = []
    for _p in data:
        _l = {}
        for _x,_y in _p.items():
            _l[_x]={"value":_y}
        _data.append(_l)
    return _data


if __name__ == "__main__":
    # 从excel读取数据
    # import_meni = [['name','姓名'],['number','电话'],['id','班级']]
    # d = ImportHander(menu=import_meni)
    # d.x = 3
    # d.y = 7
    # print public_lib.json_data(d.file(file_path='/data/temp/download/资产主机.xlsx'))



    # 导出数据到exlcel
    # w:列宽，默认为15
    # style:指定文字样式[居中/便左],默认居中
    # merge:指定合并单元格坐标[[x1,y1,x2,y2],[x11,y11,x12,y12]]
    f = {
        'menu': [['name', '姓名', {'w': 16,"style":"left"}],['number', '电话', { 'w': 6}]],
        # 'data': [
        #     {"name":{"value":"xx"},"number":{"value":"xx"}},
        #     {"name":{"value":"xx"}},
        #     {"name":{"value":"xx","background":"CD2626"}},
        # ],
        "data":[],
        'title': '学生登记表',
        'merge':[[1,4,1,10]]
    }
    # print ExportHander(**f).save('/tmp/1.xlsx')
    print ExportHander(**f).save('/tmp/1.xlsx',template={
                    "data": [],
                    "width": [{"w": 12}, {"w": 60, "style": "left"}, {"w": 12}],
                    "title": '资产项目-字段说明',
                })
    # print ExportHander(**f).response()

